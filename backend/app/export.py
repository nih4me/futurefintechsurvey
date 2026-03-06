from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
from openpyxl import Workbook

from typing import Optional

from database import get_db
from models import (
    Submission,
    ContributorInfo,
    Narrative,
    Event,
    Publication,
    GrantProject,
    Award,
    PhDStudent,
    PressAppearance,
    PartnershipProject,
    Feedback,
    PlannedContribution
)
from utils import autosize_columns

EXPORT_COLUMNS = {

    "Contributors": [
        ("submission_id", "Submission ID"),
        ("created_at", "Created At"),
        ("updated_at", "Updated At"),
        ("status", "Submission Status"),
        ("consent", "Consent Given"),
        ("email", "Email"),
        ("name", "First Name"),
        ("surname", "Last Name"),
        ("contributor_type", "Contributor Type"),
        ("affiliated_fellow_email", "Affiliated Fellow Email"),
        ("discipline", "Discipline"),
        ("has_events", "Has Events"),
        ("has_new_fundings", "Has New Fundings"),
        ("has_publications", "Has Publications"),
        ("all_publications_on_orbilu", "All Publications on ORBiLu"),
        ("has_awards", "Has Awards"),
        ("has_partnerships", "Has Partnerships"),
        ("has_phd_students", "Has PhD Students"),
        ("has_press", "Has Press Appearances"),
    ],

    "Narrative": [
        ("submission_id", "Submission ID"),
        ("narrative", "Narrative"),
    ],

    "Events": [
        ("submission_id", "Submission ID"),
        ("event_date", "Event Date"),
        ("event_name", "Event Name"),
        ("event_type", "Event Type"),
        ("location", "Location"),
        ("role", "Role"),
        ("roleComment", "Role Comment"),
    ],

    "Publications": [
        ("submission_id", "Submission ID"),
        ("publication_name", "Publication Name"),
        ("publication_date", "Publication Date"),
        ("orbilu_link", "ORBiLu Link"),
        ("mixed_gender", "Mixed Gender Team"),
        ("mixed_team", "Mixed Team"),
    ],

    "Grants": [
        ("submission_id", "Submission ID"),
        ("project_name", "Project Name"),
        ("start_date", "Start Date"),
        ("end_date", "End Date"),
        ("funder", "Funder"),
        ("funderComment", "Funder Comment"),
        ("funding_programme", "Funding Programme"),
        ("role", "Role"),
        ("roleComment", "Role Comment"),
        ("mixed_gender", "Mixed Gender Team"),
        ("mixed_team", "Mixed Team"),
    ],

    "Partnerships": [
        ("submission_id", "Submission ID"),
        ("project_name", "Project Name"),
        ("start_date", "Start Date"),
        ("partnership_type", "Partnership Type"),
        ("partner", "Partner organization"),
        ("role", "Role"),
        ("roleComment", "Role Comment"),
        ("acquired_funding", "Acquired Funding"),
    ],

    "PhD Students": [
        ("submission_id", "Submission ID"),
        ("student_name", "Student Name"),
        ("thesis_title", "Thesis Title"),
        ("graduation_year", "Graduation Year"),
        ("career_pursued", "Career Pursued"),
        ("current_work_location", "Current Work Location"),
    ],


    "Awards": [
        ("submission_id", "Submission ID"),
        ("award_date", "Award Date"),
        ("award_title", "Award Title"),
        ("award_subject", "Award Subject"),
        ("award_issuer", "Issuing Organization"),
    ],

    "Press": [
        ("submission_id", "Submission ID"),
        ("appearance_date", "Appearance Date"),
        ("press_name", "Media Outlet"),
        ("press_type", "MediaType"),
        ("appearance_type", "Appearance Type"),
        ("subject", "Subject"),
    ],

    "Planned Contributions": [
        ("submission_id", "Submission ID"),
        ("planned_text", "Planned Contributions"),
    ],

    "Feedback": [
        ("submission_id", "Submission ID"),
        ("form_intuitive", "Form Intuitiveness"),
        ("form_easy", "Ease of Information Entry"),
        ("suggestions", "Suggestions"),
    ],

}

router = APIRouter(prefix="/export", tags=["Export"])

def write_table_sheet(wb, title, columns, rows):
    ws = wb.create_sheet(title=title)

    headers = [label for field, label in columns]
    ws.append(headers)

    for row in rows:
        ws.append([getattr(row, field) for field, label in columns])

    autosize_columns(ws)

def write_contributors_sheet(wb, db: Session, submission_id=None):
    ws = wb.create_sheet(title="Contributors")

    columns = EXPORT_COLUMNS["Contributors"]

    ws.append([label for field, label in columns])

    query = (
        db.query(Submission, ContributorInfo)
        .join(
            ContributorInfo,
            ContributorInfo.submission_id == Submission.submission_id
        )
    )
    if submission_id is not None:
        query = query.filter(Submission.submission_id == submission_id)
    rows = query.all()

    for app, contributor in rows:
        data = {
            "submission_id": app.submission_id,
            "created_at": app.created_at,
            "updated_at": app.updated_at,
            "status": app.status,
            "consent": contributor.consent,
            "email": contributor.email,
            "name": contributor.name,
            "surname": contributor.surname,
            "contributor_type": contributor.contributor_type,
            "affiliated_fellow_email": contributor.affiliated_fellow_email,
            "discipline": contributor.discipline,
            "has_events": contributor.has_events,
            "has_new_fundings": contributor.has_new_fundings,
            "has_publications": contributor.has_publications,
            "all_publications_on_orbilu": contributor.all_publications_on_orbilu,
            "has_awards": contributor.has_awards,
            "has_partnerships": contributor.has_partnerships,
            "has_phd_students": contributor.has_phd_students,
            "has_press": contributor.has_press,
        }
        ws.append([data[field] for field, label in columns])

    autosize_columns(ws)

@router.get("/")
def export_submissions(submission_id: Optional[str] = None, db: Session = Depends(get_db)):
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # Contributors (custom)
    write_contributors_sheet(wb, db, submission_id)

    def get_rows(query, submission_id: Optional[str] = None):
        if submission_id is not None:
           query = query.filter_by(submission_id=submission_id)
        rows = query.all()
        return rows

    # Generic sheets
    write_table_sheet(
        wb,
        "Narrative",
        EXPORT_COLUMNS["Narrative"],
        get_rows(db.query(Narrative), submission_id)
    )

    write_table_sheet(
        wb,
        "Events",
        EXPORT_COLUMNS["Events"],
        get_rows(db.query(Event), submission_id)
    )

    write_table_sheet(
        wb,
        "Publications",
        EXPORT_COLUMNS["Publications"],
        get_rows(db.query(Publication), submission_id)
    )

    write_table_sheet(
        wb,
        "Grants",
        EXPORT_COLUMNS["Grants"],
        get_rows(db.query(GrantProject), submission_id)
    )

    write_table_sheet(
        wb,
        "Awards",
        EXPORT_COLUMNS["Awards"],
        get_rows(db.query(Award), submission_id)
    )

    write_table_sheet(
        wb,
        "PhD Students",
        EXPORT_COLUMNS["PhD Students"],
        get_rows(db.query(PhDStudent), submission_id)
    )

    write_table_sheet(
        wb,
        "Press",
        EXPORT_COLUMNS["Press"],
        get_rows(db.query(PressAppearance), submission_id)
    )

    write_table_sheet(
        wb,
        "Partnerships",
        EXPORT_COLUMNS["Partnerships"],
        get_rows(db.query(PartnershipProject), submission_id)
    )

    write_table_sheet(
        wb,
        "Partnerships",
        EXPORT_COLUMNS["Partnerships"],
        get_rows(db.query(PartnershipProject), submission_id)
    )

    write_table_sheet(
        wb,
        "Planned Contributions",
        EXPORT_COLUMNS["Planned Contributions"],
        get_rows(db.query(PlannedContribution), submission_id)
    )

    write_table_sheet(
        wb,
        "Feedback",
        EXPORT_COLUMNS["Feedback"],
        get_rows(db.query(Feedback), submission_id)
    )

    # Stream file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=submissions_export.xlsx"
        }
    )
